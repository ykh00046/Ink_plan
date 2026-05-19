import json
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook

from storage import CURRENT_FILE, SEED_FILE, create_backup, write_json_atomic


ROOT = Path(__file__).resolve().parents[1]
SOURCE = ROOT / "제품,잉크명.xlsx"
DAYS = ["월", "화", "수", "목", "금", "토", "일", "차주월"]
INK_PLAN_DAYS = ["월", "화", "수", "목", "금", "토", "일"]


def clean_cell(value):
    if value is None:
        return ""
    return str(value).strip()


def clean_ink(value):
    value = clean_cell(value)
    return None if value in ("", "-") else value


def read_products(path=SOURCE):
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb.worksheets[0]
    products = []
    machine_assignments = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        factory = clean_cell(row[0])
        name = clean_cell(row[1])
        if not name:
            continue
        inks = [clean_ink(row[3]), clean_ink(row[5]), clean_ink(row[7])]
        ink_machines = [clean_cell(row[4]), clean_cell(row[6]), clean_cell(row[8])]
        products.append({
            "factory": factory,
            "name": name,
            "type": clean_cell(row[2]),
            "brand": clean_cell(row[9]),
            "customer": clean_cell(row[9]),
            "inks": inks,
            "inkMachines": [None if m in ("", "-") else m for m in ink_machines],
        })
        for ink, machine in zip(inks, ink_machines):
            if not ink or machine in ("", "-"):
                continue
            if ink not in machine_assignments:
                machine_assignments[ink] = machine
            elif machine not in machine_assignments[ink].split(" / "):
                machine_assignments[ink] = f"{machine_assignments[ink]} / {machine}"
    return products, [
        {"ink": ink, "machine": machine}
        for ink, machine in sorted(machine_assignments.items())
    ]


def grouped_brands(products):
    groups = defaultdict(list)
    for product in products:
        brand = product.get("brand") or "미지정"
        name = product["name"]
        if name not in groups[brand]:
            groups[brand].append(name)
    return [
        {"code": brand, "label": brand, "products": names}
        for brand, names in sorted(groups.items(), key=lambda item: item[0])
    ]


def blank_schedule():
    return {day: {"day": "", "night": ""} for day in DAYS}


def reset_injection(injection):
    source = injection or {"3층": [], "1층": []}
    next_injection = {}
    for floor in ["3층", "1층"]:
        next_injection[floor] = [
            {
                "no": machine.get("no"),
                "machine": machine.get("machine") or f"{machine.get('no', '')}호기",
                "schedule": blank_schedule(),
            }
            for machine in source.get(floor, [])
        ]
    return next_injection


def blank_ink_days():
    return {
        day: {"현재고": None, "가용일수": None, "필요수량": None, "제조량": None}
        for day in INK_PLAN_DAYS
    }


def build_ink_plan(machine_assignments):
    return [
        {"name": assignment["ink"], "days": blank_ink_days()}
        for assignment in machine_assignments
    ]


def update_file(path, products, machine_assignments):
    data = json.loads(Path(path).read_text(encoding="utf-8"))
    previous_injection = data.get("injection", {})
    data["products"] = products
    data["brands"] = grouped_brands(products)
    data["machineAssignments"] = machine_assignments
    data["injectionDays"] = []
    data["injection"] = reset_injection(previous_injection)
    data["inkPlan"] = build_ink_plan(machine_assignments)
    data["inks"] = []
    data["inkAdd"] = []
    data["floor3Ink"] = []
    data["floor1Ink"] = []
    data["testInks"] = []
    data["inventory"] = {"lots": [], "daily": {}}
    write_json_atomic(path, data)
    return len(data["products"])


def main():
    products, machine_assignments = read_products()
    clean_count = update_file(SEED_FILE, products, machine_assignments)
    current_count = None
    if CURRENT_FILE.exists():
        create_backup("before_products_import")
        current_count = update_file(CURRENT_FILE, products, machine_assignments)
    print(json.dumps({
        "source": str(SOURCE),
        "products": len(products),
        "machineAssignments": len(machine_assignments),
        "clean_json": clean_count,
        "current_json": current_count,
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
